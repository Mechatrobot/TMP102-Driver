from git import Repo # Open my repository
import pandas as pd
import os
from datetime import datetime
import pytz
#from tabulate import tabulate
from styleframe import StyleFrame
from openpyxl import load_workbook
from openpyxl.styles import Alignment

url_repo = "https://github.com/Mechatrobot/TMP102-Driver"
local_file = "C:/Users/se93297/Desktop/TMP102-Driver-main"

if not os.path.exists(local_file) :
  Repo.clone_from(url_repo,local_file);       #clone the repo into the local if it doesn't exist.
else :
  #repo = Repo(local_file)
  origin = Repo(local_file).remotes.origin #remote the repo to local directory
  origin.pull()   # pull the changes into the local repository
commits =  list(Repo(local_file).iter_commits('main'))
#print(commits)
commit_data = []
for commit in commits :
   timezonedate = pytz.timezone("Africa/Casablanca")
   commit_date = commit.committed_datetime.astimezone(timezonedate)#convert the timezone to UTC.
   commit_data.append({
                       'Filename' : Repo(local_file).git.show(commit, name_only=True, format="%n"),
                       'Commiter' : commit.author.name,
                       'Email': commit.author.email,
                       'SHA' : commit.hexsha,
                       'Date of committement' : commit_date.strftime('%Y-%m-%d %H:%M:%S'), #string format
                       'Message' : commit.message.strip()
                      })
data_frame = pd.DataFrame(commit_data) #convert commit data to DataFrame to put it on excel
excel_file = os.path.join('C:/Users/se93297/Desktop/TMP102-Driver-main/file_test','commits_list.xlsx')  # create the excel file
data_frame.to_excel(excel_file, index=False, sheet_name= 'Commits') #assign DataFrame to excel
data_frame2 = pd.read_excel("file_test/extrait_dash.xlsx")
#data_frame2['Revision du scenario'] = data_frame2['Revision du scenario'].astype(float)
#print(data_frame2)

wb = load_workbook("/content/C:/Users/se93297/Desktop/TMP102-Driver-main/file_test/extrait_dash.xlsx")
sheet = wb.active

filename_extract = [cell.value for cell in sheet['A'][3:]]  
filename_commithistory = data_frame['Filename'].tolist()
SHA_extract = [cell.value for cell in sheet['D'][3:]] 
SHA_commithistory = data_frame['SHA'].tolist()
#Revision_scenario = data_frame2['Unnamed: 1'].tolist()
print(filename_extract)
#len(SHA_extract) == len(Revision_scenario)
#print(filename_extract)
#print(filename_commithistory)
#for filename in filename_commithistory :
#    if filename in filename_extract :
#        print('good')
#print(data_frame2[data_frame2.columns[0]])
for i in range(len(filename_extract)) :
  for j in range(len(filename_commithistory)) :
     if filename_extract[i] in filename_commithistory[j] :
        #SHA_var = SHA_extract[i]
            SHA_extract[i] = SHA_extract[i] + "\n" + SHA_commithistory[j]
            sheet.cell(row=i+4, column=4, value=SHA_extract[i])
            #print(SHA_extract)
            Rev = list(sheet['B{}'.format(i+4)].value.split("\n"))
            Rev_float = [float(l) for l in Rev]
            new_Rev = Rev_float[-1]+0.1000
            Rev_float.append(round(new_Rev,1))
            Rev_string = [str(l) for l in Rev_float]
                   #print(list("\n".join(Rev_string)))
                   #print(Rev_string)
            sheet.cell(row=i+4, column=2, value="\n".join(Rev_string))
print(list(sheet['B4'].value))


#sheet.cell(row=i+4 , column=2, value=float(Rev[i]))
sheet.column_dimensions['D'].width = 43
#print(SHA_extract)
#SHA_string = SHA_extract[0]
#new_sha = SHA_string.split(";")
#print(new_sha)
       #if filename_extarct[i] in data_frame2['File'].values :
       #data_frame2.loc[data_frame2['File'] == filename_extract[i], 'Revision du scenario'] += 0.1
#data_frame2.update(data_frame)
excel_updated_file = os.path.join('C:/Users/se93297/Desktop/TMP102-Driver-main/file_test','Updated_commits.xlsx')
#data_frame2.to_excel(excel_updated_file, index=False, sheet_name= 'Commits')  # create the excel file
#StyleFrame(data_frame2).to_excel(excel_updated_file,sheet_name = 'TEST').close()

#sheet.column_dimensions['A'].width = 16

#format = wb.add_format()
#format.set_align('left')
#ws.set_column('B', 10, format)
#for letter in ['A','B'] :
#  max_width = 0
#  for index in range(1,ws.max_row + 1) :
#    if len(ws[f'{letter}{index}'].value) > max_width  :
#         max_width = len(ws[f'{letter}{index}'].value)
#         print(max_width
#  ws.column_dimensions[letter].width = max_width
#for row in range(1,ws.max_row+1):
#    for col in range(1,ws.max_column+1):
#        cell=ws.cell(row, col)
#        cell.alignment = Alignment(horizontal='left')
wb.save("/content/C:/Users/se93297/Desktop/TMP102-Driver-main/file_test/test.xlsx")
#data_frame2['Revision du scenario'] = data_frame2['Revision du scenario'].astype(float)
#data_frame2['Revision du scenario'] += 0.1
#print(data_frame2['Revision du scenario'])
#print(SHA_commithistory)




